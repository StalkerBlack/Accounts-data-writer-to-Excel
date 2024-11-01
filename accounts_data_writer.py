from openpyxl import load_workbook

# Path to Excel file and text files
file_path = r"C:\path\to\your\accounts_data.xlsx"
privates_path = r"C:\path\to\your\privates.txt"
proxies_path = r"C:\path\to\your\proxies.txt"

wb = load_workbook(file_path)
ws = wb.active

# Reading data from text files
with open(privates_path, 'r') as file:
    privates = [line.strip() for line in file]

with open(proxies_path, 'r') as file:
    proxies = [line.strip() for line in file]

# Writing data in Excel
for i in range(len(privates)):  # It is assumed that the number of private keys and proxies is the same
    ws[f'A{i + 2}'] = f'Account {i + 1}'
    ws[f'B{i + 2}'] = privates[i]
    ws[f'C{i + 2}'] = proxies[i]

wb.save(file_path)
print(f'Successfully updated {file_path} with {len(privates)} accounts.')

